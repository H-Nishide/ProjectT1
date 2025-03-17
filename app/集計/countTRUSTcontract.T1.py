#######################################################################################
#######################################################################################
#  Trust契約情報 有効件数取得モジュール                  H.Nishide
#   2024.1003 T0 SQL実行～excelファイル出力まで一貫して処理
#   2024.1003 R1 excel出力　契約一覧、販社毎件数
#   2024.1008 T1 吉田さんリクエスト対応）Subsidaryの表示やめる。集計から除外する条件追加
#               販社別＋製品別件数集計
# ######################################################################################
#######################################################################################
import sys
import mysql.connector
import pandas as pd

from datetime import datetime as dt

import openpyxl as px
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from tkinter import Tk,messagebox
from tkinter import filedialog
import tkinter as tk

# import csv
# import xlwings as xw

#################################################################################
# excelの列数、カウント数字を文字表記に変換
#################################################################################
def num2alpha(num):
    if num<=26:
        return chr(64+num)
    elif num%26==0:
        return num2alpha(num//26-1)+chr(90)
    else:
        return num2alpha(num//26)+chr(64+num%26)
#################################################################################
def MessageForefrontShowinfo(MesAlarmShowinfo):
    root = Tk()
    root.attributes('-topmost', True)
    root.withdraw()
    messagebox.showinfo('確認', MesAlarmShowinfo)

###############################################################################################
# main 処理
###############################################################################################
# 販社リスト
salesCoList = ["GPAM","GPAUS","GPCN","GPEU","GPJ","GPTW","HDKR","HDSP"]
# 更新期限別リスト 更新情報リストで使用

#処理の基準日 today=dt.now()　今月の初日
print("基準日は月初！ ")
dt_today = dt.now()
#print(dt_today)
#d_today = dt_today.date()
dt_day1 = dt_today.replace(day=1)
print(dt_day1.strftime('%Y-%m-%d'))

###############################################################################################
# MySQL DBへのアクセス

#DB選択
target = ""     #本番
#test用 target = "test" #テスト

################################################################################
# GUI　#　ウインド画面の作成
################################################################################
root=tk.Tk()
root.title("TRUST契約：件数取得モジュール")
root.geometry("350x150")

#テキストの作成
text_str=tk.StringVar()
text_str.set(dt_day1.strftime('%Y-%m-%d') + "時点で有効な契約を集計します。")
text=tk.Label(textvariable=text_str)
text.pack()


#　DB環境選択ラベル
lab = tk.Label( text = "-- MySQL.DB接続環境 --" )
lab.pack()
#ラジオボタンの作成
var=tk.IntVar()
var.set('1')
radio_0=tk.Radiobutton(value=0,variable=var,text="Staging")
radio_0.pack()
radio_1=tk.Radiobutton(value=1,variable=var,text="本番")
radio_1.pack()

#テキストの作成
text_str=tk.StringVar()
text_str.set("接続先：Staging環境か本番環境か選んでください")
text=tk.Label(textvariable=text_str)
text.pack()

#ボタンを押したときに動作する関数
def check():
    text_check=var.get()
    #開始、終了
    global target 
    if text_check == 0:
        target = 'test'
    else:
        target = ''    
    text_str.set(str(text_check))
    #画面を閉じる
    root.destroy()


#ボタンの作成
Button=tk.Button(text="実行",command=check)
Button.pack()

#画面の表示
root.mainloop()



###############################################################################
# MySQLパラメータ
schema = "authtrustuser"

if target == "test":
    # #staging環境# 
    conn = mysql.connector.connect(
        host="gajp-trust-dev-mysql-v801.mysql.database.azure.com",
        user="localadmin",
        password="Screen@202410",
        db=schema
    )
else:
    #本番環境
    conn = mysql.connector.connect(
        host="gajp-trust-dev-mysql-st001.mysql.database.azure.com",
        user="localadmin",
        password="Screen@20240508",
        db=schema
    )

# カーソルを取得
cursor = conn.cursor()

##############################################################################################
# *１ 有効契約数カウント用【全体】　実行日が契約期間終了前＆TRUSTIｓOn　*/
select_all_data_query = "SELECT \
    p.Name as 販社,\
    s.model_name as Model,  \
    d.`ProductNumber` as Serial1, \
    m.`ContractExpirationDate` as ExpiredDate, \
    m.maintenance_contract_id as CntID,		/*debug用*/ \
    d.delivery_product_id as ProdID		/*debug用*/ \
FROM back_maintenance_contract m \
	JOIN back_delivery_product d ON m.delivery_product_id = d.delivery_product_id \
    JOIN back_trust_contract t ON m.trust_contract_id = t.trust_contract_id \
    JOIN back_customer c ON d.customer_id = c.customer_id \
    JOIN back_customer p ON c.parent_id = p.customer_id  \
    JOIN back_model s ON d.model_id = s.model_id \
WHERE m.`ContractExpirationDate` >= '" + dt_day1.strftime('%Y-%m-%d') + "' and  d.`TRUSTIsOn` = 1 /*契約終了日が 月初より未来である*/ \
    and c.Name <> 'GPAM Inkjet Innovation Center'\
    and c.Name <> '网屏数展示中心 Screen Digital Printing -Shanghai Showroom' \
    and c.Name <> 'GPEU STC' \
    and c.Name <> 'JiaQue' \
    and c.Name <> 'JiaRong Co.,LTD.' \
    and c.Name <> 'GPJ (MON-NAKA)' \
    and t.Item_code <> 'test' \
ORDER BY p.Name, s.model_name, d.`ProductNumber`, m.ContractExpirationDate "
# １）日付をセットして実行、条件が当日含むか否かで調整
#Pandasで実行結果をデータフレームへ
df = pd.read_sql(select_all_data_query, conn)
###############################################################################################
# 接続を閉じる
cursor.close()
conn.close()
###############################################################################################
# SQL結果0件だったら、通知して終わり
row, col = df.shape
if row < 1:
    MessageForefrontShowinfo("SQL抽出が、" + str(row) + "件でした。")
    sys.exit()

###############################################################################################
#データフレームをエクセルへ　xlsxで　テーブル作成
#ファイル名生成
tdatetime = dt.now()
tstr = tdatetime.strftime('%Y%m%d')
baseexcel = "TRUST契約件数集計." + tstr + ".xlsx"

filename = filedialog.asksaveasfilename(
    title = "名前を付けて保存",
    filetypes = [("Excelブック", ".xlsx") ], # ファイルフィルタ
    initialdir = "./", # 自分自身のディレクトリ
    initialfile = baseexcel,
    defaultextension =  "xlsx"
    )

print(filename)
baseexcel = filename

df.to_excel(baseexcel, header=True, index=False)  # +excel_name+'.xlsx')

#作ったエクセルを開く
wb = px.load_workbook(baseexcel)
ws = wb.active 
ws.title = "ContractList"
#行数、列数を得る
n = ws.max_row
m = ws.max_column
#列数を関数に投入！
p = num2alpha(m)
o = 'A1:%s%d' %(p,n)
#いざテーブル作成へ
tab = Table(displayName= "Table1", ref= o)
# Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws.add_table(tab)

# シート追加して、販社毎のデータ件数を集計
# wb.create_sheet(index=1, title='販社集計')
ws = wb.create_sheet(title = "Subtotal")

# ヘッダ行
a1 = ws.cell(1, 1)
b1 = ws.cell(1, 2)
a1.value = "販社"
b1.value = "販社件数"
ws.cell(row=1, column=3, value="製品")
ws.cell(row=1, column=4, value="製品件数")
# データ件数取得  salesCoList販社リストでまわす
i = 2
subtotal = 0
for s in salesCoList:
    # s = 'GPAM'
    df1 = df.query('販社 == @s')
    print(s)
    row, col = df1.shape
    ws.cell(row=i, column=1, value= s)
    ws.cell(row=i, column=2, value= row)
    subtotal = subtotal + row
    i += 1
    #販社内製品内訳
    m1 = "" #読み込んだ製品
    m2 = "" #処理中の製品
    mct = 0
    for m1 in df1['Model']:
        if mct == 0: #いっぱつめ
            ws.cell(row=i, column=3, value= m1)
            m2 = m1
            mct += 1
            print(m1)
        elif m1 == m2: #同一製品 2行目以降
            mct += 1
        else: #製品変わった。製品件数をexcel書きこして次の行へ
            ws.cell(row=i, column=4, value= mct)
            print(mct)
            i += 1
            ws.cell(row=i, column=3, value= m1)
            m2 = m1
            mct = 1
            print (m1)
    #最後の製品情報を書きこ
    ws.cell(row=i, column=4, value= mct)
    print(mct)
    i += 1
ws.cell(row=i, column=1, value= "subtotal")
ws.cell(row=i, column=2, value= subtotal)

#debug用強制終了
# sys.exit()

    
###############################################################################################
#ファイル保存
wb.save(baseexcel)

wb.close()  # ブックを閉じる
###############################################################################################

