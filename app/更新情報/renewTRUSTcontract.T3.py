#######################################################################################
#######################################################################################
#  販社別　Trust契約更新情報 取得モジュール      renewTRUSTcontract       H.Nishide
#   2024.1003 T0 　TRUSTcontractInfoT2.py からリメイク
#              生成Excel TRUST契約更新確認.date().xlsx
#               １）更新期限切れ一覧
#               ２）今月更新日案内の販社# 　1か月販社：GPJ、GPAM、GPAUS、CPCN、GPTW
#               ３）再来月更新日案内の販社# 　2か月販社：GPEU、HDKR、HDSP
# 2024.1004 T1 販社、更新案内期日を辞書オブジェクト化 更新案内販社、、期限別クエリ―作成
# 2024.1101 T2 再来月計算に不具合（年変わるとこ）あり改修　
# 2024.1119 T3 sqlalchemy利用でMySQL接続　
#              DB接続url取得モジュール(fenv_mysql.py)
#              DB接続パラメータ定義部分を外部ファイル「env_mysql.xml」に対応。   

#       12月に信恵さんフィードバック入れたのち、V1.0でリリースする。    

# ######################################################################################
#######################################################################################
import sys
import pymysql
import pandas as pd
import sqlalchemy as sa # MySQL接続モジュール

from datetime import datetime as dt
from dateutil.relativedelta import relativedelta

import openpyxl as px
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from tkinter import Tk,messagebox
from tkinter import filedialog
import tkinter as tk
import calendar
# import csv
# import xlwings as xw
import fenv_mysql


#################################################################################
# ワークシートをテーブルデザインつける
#################################################################################
def attachTable(fpath, sheet):
    #作ったエクセルを開く
    wb = px.load_workbook(fpath)
    ws = wb[sheet]
    #行数、列数を得る
    n = ws.max_row
    n = ws.max_row
    if n <= 1:
        return
    m = ws.max_column
    #列数を関数に投入！
    p = num2alpha(m)
    o = 'A1:%s%d' %(p,n)
    #いざテーブル作成へ
    sheet = sheet
    tab = Table(displayName= str(sheet), ref= o)
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    # 行の幅を変更
    ws.column_dimensions['M'].width = 16
    ws.column_dimensions['N'].width = 16
    # 保存して綴じる 
    wb.save(fpath)
    wb.close()  # excelブックを閉じる
######################################################
def attachTable2(wb,sheet):
    #作ったエクセルを開く
##    wb = px.load_workbook(fpath)
    ws = wb[sheet]
    #行数、列数を得る
    n = ws.max_row
    if n <= 1:
        return
    
    m = ws.max_column
    #列数を関数に投入！
    p = num2alpha(m)
    o = 'A1:%s%d' %(p,n)
    #いざテーブル作成へ
    # tab = Table(displayName= "Table1", ref= o)
#    tab = Table(displayName= "Table2", ref= o)
    sheet = sheet
    tab2 = Table(displayName= str(sheet) , ref= o)
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name='TableStyleMedium9', showRowStripes=True, showColumnStripes=False)
    tab2.tableStyleInfo = style
    ws.add_table(tab2)

##    wb.close()  # excelブックを閉じる
##    wb.save(fpath)

#################################################################################
# クエリ―実行
#################################################################################
def getContracts(pd, conn, sql_part):
    select_query = "SELECT \
        p.Name as 販社,\
        coalesce(c.SalesCompanyDivision,'NULL') as Subsidiary, \
        s.model_name as Model, \
        d.`ProductNumber` as Serial1, \
        m.`Name` as TRUST契約番号, /*_maintenance_contract,Name */ \
        c.`CustomerCode` as CustomerCode, \
        c.`Name` as Customer, /*アカウントを外す   u.`Name` as UserAccount,	*/ \
        d.`ProductNumber2` as Serial2, \
        d.`EQUIOS_dongle` as Dongle, \
        t. `Item_code` as Code, \
        t.`Name` as Name_EN, \
        t.`Name_JP` as Name_JP, \
        m.`ContractExpirationDate` as ExpiredDate, \
        m.`Valid_Until` as ValidDate, \
        m.maintenance_contract_id as ID \
    FROM back_maintenance_contract m \
        JOIN back_delivery_product d ON m.delivery_product_id = d.delivery_product_id \
        JOIN back_trust_contract t ON m.trust_contract_id = t.trust_contract_id \
        JOIN back_customer c ON d.customer_id = c.customer_id \
        JOIN back_customer p ON c.parent_id = p.customer_id \
        JOIN back_model s ON d.model_id = s.model_id " \
    + sql_part + \
    "ORDER BY p.Name, c.SalesCompanyDivision, s.model_name, d.ProductNumber, m.ContractExpirationDate, c.customer_id;"

    ###############################################################################################
    #Pandasで実行結果をデータフレームへ
    return pd.read_sql(select_query, conn)
       
 #   return dfn
#################################################################################
# 月末日を取得
#################################################################################
def get_last_date(dt):
    return dt.replace(day=calendar.monthrange(dt.year, dt.month)[1])
# 月末日を取得
#################################################################################
def get_next1_date(dt):
    print(dt.year)
    print(dt.month)
    iY = dt.year
    iM = dt.month
 #   print(dt2)
 #   return dt2.replace(day=calendar.monthrange(dt2.year, dt2.month)[1])
    return dt.replace(day=calendar.monthrange(dt.year, dt.month)[1])
#
#################################################################################
# excelの列数、カウント数字を文字表記に変換
#################################################################################
def num2alpha(num):
    if num<=26:
        return chr(64+num)
    elif num%26==0:
        return num2alpha(num//26-1)+chr(90)
    else:
        return num2alpha(num//26)+chr(64+num%26)
#################################################################################
def MessageForefrontShowinfo(MesAlarmShowinfo):
    root = Tk()
    root.attributes('-topmost', True)
    root.withdraw()
    messagebox.showinfo('確認', MesAlarmShowinfo)

###############################################################################################
# main 処理
###############################################################################################
# 販社リスト
salesCoList = ["GPAM","GPAUS","GPCN","GPEU","GPJ","GPTW","HDKR","HDSP"]
# 更新期限別リスト 更新情報リストで使用
# 　1か月販社：GPJ、GPAM、GPAUS、CPCN、GPTW
salesCoList1 = ["GPJ","GPAM","GPAUS","GPCN","GPTW"]
# 　2か月販社=GPEU、HDKR、HDSP
salesCoList2 = ["GPEU","HDKR","HDSP"]

# 販社集計期限dictionary
dCoLmt = {'GPJ': 1,'GPAM': 1,'GPAUS': 1,'GPCN': 1,'GPTW': 1,'GPEU': 2,'HDKR': 2,'HDSP': 2}
keys = dCoLmt.keys()
print(keys)  # dict_keys(['apple', 'banana', 'cherry'])
print(dCoLmt['GPAM'])
for l in keys:
    print(l)
    print(dCoLmt[l])
    


#処理の基準日
iY = dt.today().year
iM = dt.today().month
iD = dt.today().day
d = dt(iY, iM, iD)
print("今日 ")
print(d) 


#今月初
this1 = dt(iY, iM, 1)
print("今月初日")
print(this1.strftime('%Y-%m-%d'))

print("今月末 ")
# print(get_last_date(d))
thisM = get_last_date(d)
print(thisM.strftime('%Y-%m-%d'))

#// T1
#// iM += 2   #11月に　2足すと13月になるのでNG。現在値に2か月足す処理に
#// d = dt(iY, iM, iD)
#// print("再来月末 ")
#// nextM = get_last_date(d)
#// print(nextM.strftime('%Y-%m-%d'))

# T2 改修
d = thisM  + relativedelta(months=2)
print("再来月末 ")
nextM = get_last_date(d)
print(nextM.strftime('%Y-%m-%d'))

###############################################################################################
# MySQL DBへのアクセス

#DB選択
target = ""     #本番
#test用 target = "" #テスト

################################################################################
# GUI　#　ウインド画面の作成
################################################################################
root=tk.Tk()
root.title("TRUST契約更新情報 取得モジュール")
root.geometry("350x150")

#　時刻表示ラベル
lab = tk.Label( text = "-- MySQL.DB接続環境 --" )
lab.pack()

#ラジオボタンの作成
var=tk.IntVar()
var.set('1')
radio_0=tk.Radiobutton(value=0,variable=var,text="Staging")
radio_0.pack()
radio_1=tk.Radiobutton(value=1,variable=var,text="本番")
radio_1.pack()

#テキストの作成
text_str=tk.StringVar()
text_str.set("接続先：Staging環境か本番環境か選んでください")
text=tk.Label(textvariable=text_str)
text.pack()

#ボタンを押したときに動作する関数
def check():
    text_check=var.get()
    #開始、終了
    global target 
    if text_check == 0:
        target = 'stag'
    else:
        target = 'prod'
    text_str.set(str(text_check))
    #画面を閉じる
    root.destroy()

#ボタンの作成
Button=tk.Button(text="実行",command=check)
Button.pack()

#画面の表示
root.mainloop()

###############################################################################
###############################################################################
# MySQL接続
# 接続先取得 # url = f'mysql+pymysql://{user}:{password}@{host}:{port}/{db}?charset=utf8'
url = fenv_mysql.getURL(target)
if url == "":
    print("★★★　MySQL　接続URL取得エラー　")
    sys.exit()

# secure_transport=ON との接続　呪文みたいなもん
connect_args={'ssl':{'fake_flag_to_enable_tls': True}}
try:
    # engine作成
    # engine = sa.create_engine(url, echo=False)
    engine = sa.create_engine(url, connect_args=connect_args)
except sa.create_engine as e:
     print(f'★★★★ Error create_engine: {e}')
     sys.exit()
##############################################################################################
# データ取得のクエリ
# /*　契約更新日過ぎてるぞリストの取得　*/

select_all_data_query = "SELECT \
    p.Name as 販社,\
    coalesce(c.SalesCompanyDivision,'NULL') as Subsidiary, /*販社と分けて出す*/ \
    s.model_name as Model, \
    d.`ProductNumber` as Serial1, \
    m.`Name` as TRUST契約番号, 		/*追加）トラスト契約番号 _maintenance_contract,Name */ \
    c.`CustomerCode` as CustomerCode, \
    c.`Name` as Customer, 	/*アカウントを外す   u.`Name` as UserAccount,	*/ \
    d.`ProductNumber2` as Serial2, \
    d.`EQUIOS_dongle` as Dongle, \
    t. `Item_code` as Code, \
    t.`Name` as Name_EN, \
    t.`Name_JP` as Name_JP, \
    m.`ContractExpirationDate` as ExpiredDate, \
    m.`Valid_Until` as ValidDate, \
    m.maintenance_contract_id as ID \
FROM back_maintenance_contract m \
	JOIN back_delivery_product d ON m.delivery_product_id = d.delivery_product_id \
    JOIN back_trust_contract t ON m.trust_contract_id = t.trust_contract_id \
    JOIN back_customer c ON d.customer_id = c.customer_id \
    JOIN back_customer p ON c.parent_id = p.customer_id \
    JOIN back_model s ON d.model_id = s.model_id \
WHERE m.`ContractExpirationDate` < '" + this1.strftime('%Y-%m-%d') + "' and  d.`TRUSTIsOn` = 1 /*契約終了日が実行日より未来である*/ \
ORDER BY p.Name, c.SalesCompanyDivision, s.model_name, d.`ProductNumber`, m.ContractExpirationDate, c.customer_id	/*販社順,期限日順*/"

###############################################################################################
#Pandasで実行結果をデータフレームへ# df = pd.read_sql(select_all_data_query, conn)
#Pandasで実行結果をデータフレームへ https://qiita.com/ryo19841204/items/811016a4f09489e58eb1
df = pd.read_sql(select_all_data_query, con=engine)
    
###############################################################################################
# SQL結果0件だったら、通知して終わり
row, col = df.shape
if row < 1:
    MessageForefrontShowinfo("SQL抽出が、" + str(row) + "件でした。")
    sys.exit()

###############################################################################################
#データフレームをエクセルへ　xlsxで　テーブル作成
#ファイル名生成
tdatetime = dt.now()
tstr = tdatetime.strftime('%Y%m%d')
baseexcel = "TRUST契約更新確認." + tstr + ".xlsx"

filename = filedialog.asksaveasfilename(
    title = "名前を付けて保存",
    filetypes = [("Excelブック", ".xlsx") ], # ファイルフィルタ
    initialdir = "./", # 自分自身のディレクトリ
    initialfile = baseexcel,
    defaultextension =  "xlsx"
    )

print(filename)
baseexcel = filename

# df.to_excel(baseexcel, header=True, index=False, sheet_name="期限切れリスト") 
df.to_excel(baseexcel, header=True, index=False, sheet_name="expieredList") 

#作ったエクセルを開く
wb = px.load_workbook(baseexcel)
ws = wb.active 
#行数、列数を得る
n = ws.max_row
m = ws.max_column
#列数を関数に投入！
p = num2alpha(m)
o = 'A1:%s%d' %(p,n)
#いざテーブル作成へ
tab = Table(displayName= "Table1", ref= o)
# Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws.add_table(tab)
# 行の幅を変更
ws.column_dimensions['M'].width = 16
ws.column_dimensions['N'].width = 16
# 保存して綴じる 
wb.save(baseexcel)
wb.close()  # excelブックを閉じる

###############################################################################################
# 販社毎更新案内シート生成ループ
# 1か月案内
# 2か月案内
# シート追加して、販社毎のデータ件数を集計
# wb.create_sheet(index=1, title='販社集計')

# writer_spec = {"path":"既存エクセル.xlsx",
writer_spec = { "path":baseexcel ,
               "engine":"openpyxl",
               "mode":"a",
               "if_sheet_exists":"overlay"}

    
# データ件数取得  salesCoList販社リストでまわす
keys = dCoLmt.keys()
print(keys)  # dict_keys(['apple', 'banana', 'cherry'])
print(dCoLmt['GPAM'])
for s1 in keys:
    print(s1)    #販社
    print(dCoLmt[s1])    #期限
    # SQL実行
    # PandasとWHERER句（販社、BETWEENのstart,end）を引数に関数化、返値がデータフレーム
    if dCoLmt[s1] == 1:  #今月末までリストの販社
        sql_where_part = "WHERE p.Name = '" + s1 + "' and  d.`TRUSTIsOn` = 1 and m.ContractExpirationDate BETWEEN '" + this1.strftime('%Y-%m-%d') + "' and '" + thisM.strftime('%Y-%m-%d') + "'"
    elif dCoLmt[s1] == 2:  #来月末までリストの販社
        sql_where_part = "WHERE p.Name = '" + s1 + "' and  d.`TRUSTIsOn` = 1 and m.ContractExpirationDate BETWEEN '" + this1.strftime('%Y-%m-%d') + "' and '" + nextM.strftime('%Y-%m-%d') + "'"
    else:
        continue
            
    df1 = getContracts(pd, engine, sql_where_part)
    # データフレームをExcelに流し込み
    with pd.ExcelWriter(**writer_spec) as writer:
        df1.to_excel(writer,  header=True, index=False, sheet_name=s1)  
    
    attachTable(baseexcel, s1)   # Excel開いてテーブル化して閉じる
#★    attachTable(baseexcel, s1) を使わずに　もうひとループここでやる
#★wb = px.load_workbook(baseexcel)
#★for s1 in keys:
#★        attachTable2(wb, s1)
#★wb.save(baseexcel)
#★wb.close()  # ブックを閉じる


###############################################################################################
# wb.save(baseexcel)
# wb.close()  # ブックを閉じる
###############################################################################################

###############################################################################################
# DB接続を閉じる
# cursor.close()
# conn.close()
###############################################################################################
